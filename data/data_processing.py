import warnings
warnings.filterwarnings("ignore", "\nPyarrow", DeprecationWarning)
warnings.simplefilter("ignore")
 
from openpyxl import load_workbook
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QRadioButton, QLineEdit, QPushButton, QMessageBox, QGridLayout, QFileDialog, QComboBox, QVBoxLayout, QMainWindow, QProgressBar
from PyQt5.QtGui import QPixmap, QIcon
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QThread
from playwright._impl._errors import TargetClosedError

import sys
import os
import math
import unicodedata
import re
import shutil
import pandas as pd
import pyautogui
import subprocess
import time

############################################################
####################### ALERTAS ############################
############################################################

def show_error_message(title, message):
    app = QApplication.processEvents()
    error_box = QMessageBox()
    error_box.setWindowIcon(QIcon(os.getcwd() + "/assets/ntt_icone.ico"))
    error_box.setIcon(QMessageBox.Critical)
    error_box.setWindowTitle(title)
    error_box.setText(message)
    error_box.setStandardButtons(QMessageBox.Ok)
    error_box.exec_()
    QApplication.processEvents()

def show_warning_message(title, message):
    app = QApplication.processEvents()
    warning_box = QMessageBox()
    warning_box.setWindowIcon(QIcon(os.getcwd() + "/assets/ntt_icone.ico"))
    warning_box.setIcon(QMessageBox.Warning)
    warning_box.setWindowTitle(title)
    warning_box.setText(message)
    warning_box.setStandardButtons(QMessageBox.Ok)
    warning_box.exec_()
    QApplication.processEvents()

def show_information_message(title, message):
    app = QApplication.processEvents()
    information = QMessageBox()
    information.setWindowIcon(QIcon(os.getcwd() + "/assets/ntt_icone.ico"))
    information.setIcon(QMessageBox.Information)
    information.setWindowTitle(title)
    information.setText(message)
    information.setStandardButtons(QMessageBox.Ok)
    information.exec_()
    QApplication.processEvents()

def show_information_message_with_link(title, message, link):
    msg_box = QMessageBox()
    msg_box.setWindowIcon(QIcon(os.getcwd() + "/assets/ntt_icone.ico"))
    msg_box.setIcon(QMessageBox.Information)
    msg_box.setWindowTitle(title)
    msg_box.setText(message)

    # Adiciona um bot√£o com o texto "Copiar link"
    copy_button = msg_box.addButton("Copiar link", QMessageBox.ActionRole)
    copy_button.clicked.connect(lambda: QApplication.clipboard().setText(link))

    msg_box.exec_()
    
############################################################
############## EXPORTA√á√ÉO DE DADOS DO BROWSER ##############
############################################################

if not os.path.exists("browser"):
    # Nome do arquivo autoextra√≠vel
    arquivo_autoextraivel = "browser.exe"

    # Executar o arquivo autoextra√≠vel
    processo = subprocess.Popen(arquivo_autoextraivel)

    # Esperar um segundo para garantir que a janela apare√ßa
    time.sleep(1)

    # Clicar diretamente no bot√£o "Extract"
    pyautogui.press('enter')

    # Minimizar a janela ativa usando a combina√ß√£o de teclas Win + ‚Üì
    pyautogui.hotkey('win', 'down')

    # Esperar a extra√ß√£o finalizar
    processo.wait()
 
############################################################
################ INTERFACE DO USU√ÅRIO ######################
############################################################

# Vari√°vel global para rastrear qual bot√£o foi clicado
botao_clicado = None

def selecionar_so_calculadora_sigla():
    global sigla
    # global tipo_calculadora_selecionado
    global so_selecionado
    global botao_clicado

    # Obtendo o texto do bot√£o de r√°dio selecionado
    if radio_button_1.isChecked():
        so_selecionado = radio_button_1.text()
    elif radio_button_2.isChecked():
        so_selecionado = radio_button_2.text()
    elif radio_button_3.isChecked():
        so_selecionado = radio_button_3.text()
    else:
        QMessageBox.warning(root, "Aviso", "Voc√™ precisa selecionar um sistema operacional")
        return
    
    # Obtendo o texto selecionado do combobox
    # tipo_calculadora_selecionado = combobox.currentText()
    
    sigla = comentario_entry.text().upper()
    if sigla:
        botao_clicado = "Confirmar"
        # if tipo_calculadora_selecionado == "Calculadora To Be":
        #     QMessageBox.warning(root, "Aviso", "A op√ß√£o 'Calculadora To Be' est√° em desenvolvimento")
        #     botao_clicado = "To Be"

        print("\n ------------------------------------------------------------ \n")
        print("Iniciando automa√ß√£o")
        print("\nObtendo Sistema Operacional: " + so_selecionado)
        # print("\nObtendo Tipo de Calculadora: " + tipo_calculadora_selecionado)
        print("\nObtendo informa√ß√µes da sigla: " + sigla)

        root.close()
    else:
        QMessageBox.warning(root, "Aviso", "Voc√™ precisa preencher uma sigla")

def on_close(event):
    # A√ß√£o personalizada ao fechar a janela
    if botao_clicado == "To Be" or botao_clicado is None:
        reply = QMessageBox()
        reply.setWindowTitle('Aviso')
        reply.setWindowIcon(QIcon(os.getcwd() + "/assets/ntt_icone.ico"))
        reply.setText('Tem certeza que deseja fechar?')
        reply.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        reply.button(QMessageBox.Yes).setText('Sim')
        reply.button(QMessageBox.No).setText('N√£o')
        reply.setIcon(QMessageBox.Question)

        if reply.exec() == QMessageBox.Yes:
            exit(1)
        else:
            event.ignore()
 
def ajustar_dimensionamento():
    largura_janela = int(app.primaryScreen().size().width() * 0.4) # 40% da largura da tela
    altura_janela = int(app.primaryScreen().size().height() * 0.4) # 40% da altura da tela
    return largura_janela, altura_janela
 
app = QApplication(sys.argv)
root = QWidget()

root.setWindowTitle("Script Calculadora AWS")
root.setWindowIcon(QIcon(os.getcwd() + "/assets/ntt_icone.ico"))
 
largura_janela, altura_janela = ajustar_dimensionamento()
root.resize(largura_janela, altura_janela)
root.setFixedSize(((app.primaryScreen().size().width() - largura_janela) // 2) + 100,
          ((app.primaryScreen().size().height() - altura_janela) // 2) + 100)
root.move((app.primaryScreen().size().width() - largura_janela) // 2,
          (app.primaryScreen().size().height() - altura_janela) // 2)
 
# Logo da empresa
logo_label = QLabel(root)
pixmap = QPixmap(os.getcwd() + '/assets/ntt_logo.png')
max_width_logo = int(largura_janela * 0.4)  # Definindo uma largura m√°xima para a logo
pixmap = pixmap.scaledToWidth(max_width_logo)
logo_label.setPixmap(pixmap)
 
layout = QGridLayout(root)
layout.addWidget(logo_label, 0, 0, 1, 2, alignment=Qt.AlignHCenter)

# Label abaixo do logo
label_informativa = QLabel("Selecione as op√ß√µes de S.O, tipo de calculadora e preencha a sigla abaixo:")
label_informativa.setStyleSheet("font-size: 14px")
layout.addWidget(label_informativa, 1, 0, 1, 2, alignment=Qt.AlignCenter)

# # Lista de radio buttons
opcoes = ["Windows", "Linux", "Windows e Linux"]

# Criando um QVBoxLayout para os radio buttons
radio_button_layout = QVBoxLayout()

# Adicionando os radio buttons ao layout
radio_button_1 = QRadioButton(opcoes[0])
radio_button_1.setStyleSheet("font-size: 18px; margin-left: 265px; margin-top: 20px;")
radio_button_layout.addWidget(radio_button_1)

radio_button_2 = QRadioButton(opcoes[1])
radio_button_2.setStyleSheet("font-size: 18px; margin-left: 265px; margin-top: 20px;")
radio_button_layout.addWidget(radio_button_2)

radio_button_3 = QRadioButton(opcoes[2])
radio_button_3.setStyleSheet("font-size: 18px; margin-left: 265px; margin-top: 20px;")
radio_button_layout.addWidget(radio_button_3)

# Adicionando espa√ßo vazio
radio_button_layout.addStretch()

# Adicionando o layout de radio buttons ao layout principal
layout.addLayout(radio_button_layout, 4, 0)

# # # Criando um QVBoxLayout para o combobox
# combobox_layout = QVBoxLayout()
# combobox_layout.addStretch() # Adicionando espa√ßo vazio no layout para ajustar a altura do combobox
 
# # Adicionando QComboBox ao lado da lista de radio buttons
# combobox = QComboBox()
# combobox.addItem("Calculadora MAP")
# combobox.addItem("Calculadora To Be")
# combobox.setStyleSheet("font-size: 16px;")
# combobox_layout.addWidget(combobox)
# combobox_layout.addStretch() # Adicionando espa√ßo vazio no layout
# combobox_layout.addStretch() # Adicionando espa√ßo vazio no layout
# combobox_widget = QWidget()
# combobox_widget.setLayout(combobox_layout)
# layout.addWidget(combobox_widget, 2, 1, len(opcoes), 1, alignment=Qt.AlignLeft)

# Criando um QVBoxLayout para o campo de entrada de texto
entry_layout = QVBoxLayout()

# Adicionando espa√ßo vazio
entry_layout.addStretch()

# Campo de entrada de texto
comentario_entry = QLineEdit(root)
comentario_entry.setPlaceholderText("Digite a sigla")
comentario_entry.setStyleSheet("width: 110px; height: 30px; font-size: 16px; border-radius: 5px;")
layout.addWidget(comentario_entry, len(opcoes) + 2, 0, 1, 2, alignment=Qt.AlignCenter)

# Bot√£o de confirma√ß√£o
confirmar_button = QPushButton('Confirmar', root)
confirmar_button.clicked.connect(selecionar_so_calculadora_sigla)
confirmar_button.setStyleSheet("background-color: #007BC4; color: white; width: 110px; height: 30px; font-size: 16px; border-radius: 5px;")
layout.addWidget(confirmar_button, len(opcoes) + 3, 0, 1, 2, alignment=Qt.AlignCenter)

# Assinatura com links para o LinkedIn
assinatura_label = QLabel()
assinatura_text = "Develop by: <a href=\"https://www.linkedin.com/in/anderson-castro-ribeiro-34b192114/\" style=\"color:blue;\">Anderson</a> e <a href=\"https://www.linkedin.com/in/mrk-silva/\" style=\"color:blue;\">Marco</a>"
assinatura_label.setOpenExternalLinks(True)  # Para abrir o link em um navegador externo
assinatura_label.setText(assinatura_text)
assinatura_label.setStyleSheet("font-size: 14px;")
assinatura_label.setTextInteractionFlags(Qt.TextBrowserInteraction)
layout.addWidget(assinatura_label, len(opcoes) + 4, 0, 1, 2, alignment=Qt.AlignCenter)

# Label de ajuda
ajuda_label = QLabel("Ajuda", root)
ajuda_label.setStyleSheet("font-size: 12px; color: white; background-color: #007BC4; padding: 5px; border-radius: 5px;")
ajuda_label.setAlignment(Qt.AlignRight | Qt.AlignBottom)
ajuda_label.setCursor(Qt.PointingHandCursor)

def on_help_clicked():
    help_window = QMessageBox()
    help_window.setWindowTitle("Instru√ß√µes")
    help_window.setWindowIcon(QIcon(os.getcwd() + "/assets/ntt_icone.ico"))
    help_window.setIcon(QMessageBox.Information)

    # Mensagem principal em fonte maior
    main_message = "<h2>Passo a passo para a utiliza√ß√£o do script</h2>"
    main_message += "<p style='font-size: 14px; color: gray'> <span style='color: red;'><b>Aten√ß√£o!</b></span> Antes de executar o script, certifique-se de que o diret√≥rio <b>N√ÉO SEJA</b> dentro do <span style='color: blue;'>Microsoft OneDrive</span></p>"
    main_message += "<br><p style='font-size: 14px;'>1 - Selecione o sistema operacional que queira filtrar da sigla</p>"
    # main_message += "<p style='font-size: 14px;'>2 - Selecione o tipo de calculadora</p>"
    main_message += "<p style='font-size: 14px;'>2 - Preencha a sigla</p>"
    main_message += "<p style='font-size: 14px;'>3 - Clique em 'confirmar'</p>"
    main_message += "<p style='font-size: 14px;'>4 - A seguir ir√° abrir a sele√ß√£o de arquivos, voc√™ deve selecionar o arquivo de dados contendo os servidores</p>"
    main_message += "<p style='font-size: 14px;'>5 - Ser√£o exibidas as informa√ß√µes sobre os dados que ser√£o tratados</p>"
    main_message += "<p style='font-size: 14px;'>6 - Iniciar√° a automa√ß√£o utilizando o navegador (lembre-se de deixar o navegador em foco na tela at√© finalizar a automa√ß√£o)</p>"
    main_message += "<p style='font-size: 14px;'>7 - Ao final, clique no bot√£o para copiar o link da calculadora gerada</p>"

    # Adicionando uma linha extra de espa√ßo
    main_message += "<br>"

    # Mensagem adicional em fonte menor
    additional_message = "<p style='font-size: 14px;'>Para mais detalhes, consulte os desenvolvedores üíª</p>"

    # Combine as mensagens
    full_message = main_message + additional_message

    help_window.setText(full_message)
    help_window.setStandardButtons(QMessageBox.Ok)
    help_window.exec_()

ajuda_label.mousePressEvent = lambda event: on_help_clicked()
layout.addWidget(ajuda_label, len(opcoes) + 4, 1, alignment=Qt.AlignRight | Qt.AlignBottom)

root.setLayout(layout)

# Adicionando o evento de fechar a janela
root.closeEvent = lambda event: on_close(event)

root.show()

app.exec_()

############################################################
################# TRATAMENTO DE DADOS ######################
############################################################

def selecionar_arquivo():
    print("\n ------------------------------------------------------------ \n")
    caminho_arquivo, _ = QFileDialog.getOpenFileName(None, "Selecione o arquivo de dados", filter="Arquivos Excel (*.xlsx)")
    return caminho_arquivo

def selecionar_caminho_salvar():
    caminho_pasta = "data/processing"
    arquivo = "data_sigla.xlsx"
    caminho_arquivo = os.path.join(os.getcwd(),caminho_pasta,arquivo)
    os.remove(caminho_arquivo) if os.path.exists(caminho_arquivo) else None
    return caminho_arquivo

def validar_aba_excel(caminho_arquivo, aba):
    xls = pd.ExcelFile(caminho_arquivo)
 
    abas_disponiveis = xls.sheet_names
    # Verifica se a aba desejada est√° presente na lista
    if aba in abas_disponiveis:
        return True
    else:
        return False

def converter_utf8(caminho_arquivo, aba):
    print("Iniciando tratamento de dados. \n")
 
    # Carregar o arquivo Excel com a codifica√ß√£o ISO-8859-1
    df = pd.read_excel(caminho_arquivo, sheet_name=aba)
 
    # Iterar sobre as c√©lulas na coluna 'FUNCAO'
    for index, row in df.iterrows():
        # Obter o texto original da c√©lula 'FUNCAO'
        texto_original = row['FUNCAO']
        texto_original = str(texto_original)
 
 
        # Verificar se o texto precisa de convers√£o
        if necessita_conversao(texto_original):
            # Converter o texto e remover acentos
            texto_convertido = corrigir_e_converter_texto(texto_original)
        else:
            # Manter o valor original
            texto_convertido = texto_original
 
        texto_convertido = texto_convertido.replace(">", "")
        texto_convertido = texto_convertido.replace("<", "")
        texto_convertido = texto_convertido.replace("&", "")
 
        # Adicionar o valor tratado na coluna 'FUNCAO' para a linha atual
        df.at[index, 'FUNCAO'] = remover_acentos(texto_convertido)
 
    # Salvar o DataFrame de volta no arquivo Excel com a codifica√ß√£o UTF-8
    df.to_excel(caminho_arquivo, sheet_name=aba, index=False)

def necessita_conversao(texto):
    # Lista de caracteres em ISO-8859-1
    caracteres_iso88591 = "¬£¬°¬ß‚Ä°∆í¬¢√Ü‚Ç¨‚Ä¢"
 
    # Verificar a presen√ßa de qualquer caractere em ISO-8859-1 no texto
    for caractere in caracteres_iso88591:
        if caractere in texto:
            return True
 
    # Se nenhum caractere em ISO-8859-1 foi encontrado, n√£o √© necess√°rio converter
    return False

def corrigir_e_converter_texto(texto):
    # Corrigir caracteres espec√≠ficos
    texto = texto.replace("√É‚Ä°√É∆íO", "CAO")
    texto = texto.replace("√É¬ß√É¬£o", "CAO")
    texto = texto.replace("√É‚Ä°O", "CO")
    texto = texto.replace("√É¬ßa", "CA")
    texto = texto.replace("√É∆í", "√É¬£")
    texto = texto.replace("√¢‚Ç¨‚Äú", "")
    texto = texto.replace("  ", " ")
    texto = texto.replace("√É‚Ä¢", "O")
    texto = texto.replace("√É¬™", "E")
    texto = texto.replace("√É‚Ä∞", "E")
    texto = texto.replace("√É≈†", "E")
 
    # Remover acentos
    texto = unicodedata.normalize("NFD", texto).encode("ascii", "ignore").decode("utf-8")
 
    return texto

def remover_acentos(texto):
    # Remover acentos
    texto_sem_acento = unicodedata.normalize("NFD", texto).encode("ascii", "ignore").decode("utf-8")
 
    return ajustar_case(texto_sem_acento)

def ajustar_case(texto):
    # Quebra a string em palavras
    palavras = texto.split()
    # Inicializa a string resultante
    texto_ajustado = ""
    # Percorre cada palavra
    for palavra in palavras:
        # Verifica se a palavra est√° toda em mai√∫sculas
        if palavra != palavra.upper():
            # Ajusta o caso da palavra
            texto_ajustado += ajustar_caso_palavra(palavra) + " "
        else:
            # Se a palavra est√° toda em mai√∫sculas, mant√©m como est√°
            texto_ajustado += palavra + " "
    # Remove o √∫ltimo espa√ßo adicionado
    texto_ajustado = texto_ajustado.strip()
    return texto_ajustado

def ajustar_caso_palavra(palavra):
    # Ajusta o caso da palavra
    if palavra.islower():
        return palavra
    elif len(palavra) > 3:
        return palavra[0].upper() + palavra[1:].lower()
    elif len(palavra) == 3:
        return palavra
    else:
        return palavra.lower()

def organizar_colunas(caminho_arquivo, aba):
    # Carrega o arquivo Excel
    xls = pd.ExcelFile(caminho_arquivo)
    planilha_original = pd.read_excel(xls, sheet_name=aba)
 
    # Renomear as colunas
    planilha_original.rename(columns={'TIPO_APLICACAO': 'DESCRICAO', 'vcpu': 'CPU', 'MEMORIA_RAM': 'MEMORIA', 'STORAGE': 'STORAGE (GB)'}, inplace=True)
 
    # Reordenar as colunas conforme necess√°rio
    planilha_original = planilha_original[['NOME', 'AMBIENTE', 'DESCRICAO', 'STORAGE (GB)', 'MEMORIA', 'CPU', 'FUNCAO', 'CLASSE']]
 
    # Concatenar o texto das colunas C, E, F e G e adicionar na coluna C
    planilha_original['DESCRICAO'] = planilha_original.apply(lambda row: f"{row['DESCRICAO']} {row['MEMORIA']}GB RAM {row['CPU']}vCPU ({row['FUNCAO']})", axis=1)
 
    # Salva as altera√ß√µes no arquivo Excel
    planilha_original.to_excel(caminho_arquivo, sheet_name=aba, index=False)

def combinacoes_configuracoes(caminho_arquivo, aba):
    # Carregar a planilha Excel
    df = pd.read_excel(caminho_arquivo, sheet_name=aba)
 
    # Concatenar 'AMBIENTE', 'DESCRICAO' e 'CLASSE' em uma nova coluna 'COMBINACAO'
    df['COMBINACAO'] = df['AMBIENTE'] + ';' + df['DESCRICAO'] + ';' + df['CLASSE']
    df['STORAGE'] = df['STORAGE (GB)']
 
    # Calcular o maior "storage" para cada combina√ß√£o
    max_storage_combinacao = {}
    for _, row in df.iterrows():
        combinacao = row['COMBINACAO']
        storage = row['STORAGE']
        if combinacao not in max_storage_combinacao:
            max_storage_combinacao[combinacao] = storage
        else:
            max_storage_combinacao[combinacao] = max(max_storage_combinacao[combinacao], storage)
 
    # Contar o n√∫mero de ocorr√™ncias de cada combina√ß√£o
    contagem_combinacoes = df['COMBINACAO'].value_counts().reset_index()
    contagem_combinacoes.columns = ['COMBINACAO', 'QUANTIDADE']
 
    # Formatar os resultados no formato desejado e aplica calculo de Multi-AZ e Blue/Green
    def formatar_resultado(row):
        combinacao = row['COMBINACAO']
        quantidade = row['QUANTIDADE']
        partes = combinacao.split(';')
        max_storage = max_storage_combinacao.get(combinacao, 0)
        if len(partes) >= 3:
            ambiente = partes[0]
            descricao = partes[1]
            classe = partes[2]
            web = "WEB SERVERS"
 
            if ambiente.upper() == "DEVELOPMENT" or ambiente.upper() == "DESENVOLVIMENTO" or ambiente.upper() == "HOMOLOGATION" or ambiente.upper() == "HOMOLOGACAO" or ambiente.upper() == "PRE-PRODUCTION" or ambiente.upper() == "PRE-PRODUCAO":
                if web in descricao:
                    return f"{ambiente};{quantidade*4};{quantidade*4} x {descricao};{classe};{max_storage}"
                else:
                    return f"{ambiente};{quantidade*2};{quantidade*2} x {descricao};{classe};{max_storage}"
 
            elif ambiente.upper() == "PRODUCTION" or ambiente.upper() == "PRODUCAO":
                if web in descricao:
                    return f"{ambiente};{quantidade*6};{quantidade*6} x {descricao};{classe};{max_storage}"
                else:
                    return f"{ambiente};{quantidade*3};{quantidade*3} x {descricao};{classe};{max_storage}"
            else:
                return f"{ambiente};{quantidade};{quantidade} x {descricao};{classe};{max_storage}"
        else:
            return f"{combinacao};{quantidade};{max_storage}"
 
    contagem_combinacoes['RESULTADO'] = contagem_combinacoes.apply(formatar_resultado, axis=1)
 
    return contagem_combinacoes['RESULTADO']

def criar_aba_com_resultados(caminho_arquivo, resultados):
    workbook = load_workbook(caminho_arquivo)
    aba = 'Controle'
 
    nova_aba = workbook.create_sheet(aba)
 
    aba_controle = workbook[aba]
 
    mapeamento_colunas = {
        "Group": 0, "Description": 1, "AWS Region": 2, "Operating System": 3, "Instance Type": 4,
        "Tenancy": 5, "Number of Instances": 6, "Assumed Usage": 7, "Usage Type": 8, "Purchasing Option": 9,
        "Storage Type": 10, "Storage amount per Instance (GB)": 11,"Provisioning IOPS per instance (applicable for gp3, io1, io2)": 12, "EBS Throughput per Instance (applicable for gp3)(Mbps)": 13,
        "Snapshot Frequency": 14, "EBS Snapshot amount per Instance (GB/snapshot)": 15
    }
 
    # Adicionar cabe√ßalhos
    cabecalhos = list(mapeamento_colunas.keys())
    nova_aba.append(cabecalhos)
 
    ambiente_adicional = []
    # Preencher a nova aba com os resultados
    for resultado in resultados:
        if resultado is not None: #Verifica se o resultado n√£o √© None
            partes = resultado.split(';')
            descricao = '"' + partes[2] + '"'
 
            if partes[3] == "Linux Server":
                partes[3] = "Linux"
 
            if partes[0].upper() == "DEVELOPMENT" or partes[0].upper() == "DESENVOLVIMENTO" or partes[0].upper() == "HOMOLOGATION" or partes[0].upper() == "HOMOLOGACAO" or partes[0].upper() == "PRE-PRODUCTION" or partes[0].upper() == "PRE-PRODUCAO":
                linha = [partes[0], descricao,'sa-east-1', partes[3], '', 'Shared Instances', partes[1], '40', 'Hours/Week', 'On-Demand', 'General Purpose SSD (gp3)', partes[4], '', '', '2x Daily', '10']
            elif partes[0].upper() == "PRODUCTION" or partes[0].upper() == "PRODUCAO":
                linha = [partes[0], descricao,'sa-east-1', partes[3], '', 'Shared Instances', partes[1], '', 'Always On', '1 Yr No Upfront EC2 Instance Savings Plan', 'General Purpose SSD (gp3)', partes[4], '', '', '6x Daily', '20']
            else: # Outros ambientes
                linha = [partes[0], descricao,'sa-east-1', partes[3], '', 'Shared Instances', partes[1], '40', 'Hours/Week', 'On-Demand', 'General Purpose SSD (gp3)', partes[4], '', '', '2x Daily', '10']
                ambiente_adicional.append(partes[0])
            nova_aba.append(linha)

    environment_add = False
 
    if ambiente_adicional:
        ambiente_adicional = list(set(ambiente_adicional))
        ambiente_adicional = ["Disaster Recovery (DR)" if str(ambiente) == "DR" else str(ambiente) for ambiente in ambiente_adicional]
        ambiente_adicional = ', '.join(ambiente_adicional)

        environment_add = True

    if environment_add == True:
        default_message = "<p style='font-size: 14px;'>Para cada ambiente foram aplicadas as regras abaixo:"
        default_message += "<p style='font-size: 14px;'><b>Ambiente com WEBSERVERS (Multi AZ e Blue/Green):</b>"
        default_message += "<p style='font-size: 14px;'>Development = 2 AZs x 2 Blue/Green<br>Homologation = 2 AZs x 2 Blue/Green<br>Production = 3 AZs x 2 Blue/Green"
        default_message += "<p style='font-size: 14px;'><b>Ambiente sem WEBSERVERS (Multi AZ):</b>"
        default_message += "<p style='font-size: 14px;'>Development = 2 AZs<br>Homologation = 2 AZs<br>Production = 3 AZs"
        default_message += "<p style='font-size: 14px;'><b>N√£o foram adicionados servidores de </b><b style='color: red;'>BANCO DE DADOS</b>"
        default_message += "<br><p style='font-size: 14px; color: red; font-weight: bold;'>Aten√ß√£o!"
        default_message += "<p style='font-size: 14px;'>Para essa sigla, foram encontrados ambientes fora do padr√£o (Development, Homologation e Production), abaixo ser√£o listados os ambientes:"
        default_message += f"<p style='font-size: 14px; color: red; font-weight: bold;'>{ambiente_adicional}"
        show_information_message("Aviso", default_message)

    else:
        default_message = "<p style='font-size: 14px;'>Para cada ambiente foram aplicadas as regras abaixo:"
        default_message += "<p style='font-size: 14px;'><b>Ambiente com WEBSERVERS (Multi AZ e Blue/Green):</b>"
        default_message += "<p style='font-size: 14px;'>Development = 2 AZs x 2 Blue/Green<br>Homologation = 2 AZs x 2 Blue/Green<br>Production = 3 AZs x 2 Blue/Green"
        default_message += "<p style='font-size: 14px;'><b>Ambiente sem WEBSERVERS (Multi AZ):</b>"
        default_message += "<p style='font-size: 14px;'>Development = 2 AZs<br>Homologation = 2 AZs<br>Production = 3 AZs"
        default_message += "<p style='font-size: 14px;'><b>N√£o foram adicionados servidores de </b><b style='color: red;'>BANCO DE DADOS</b>"
        show_information_message("Aviso", default_message)
    
    # Iterar sobre as c√©lulas da coluna B (Descricao) e remover as aspas
    for linha in aba_controle.iter_rows(min_row=2, min_col=2, max_col=2):  # Coluna B (Descricao) come√ßando da segunda linha
        for cell in linha:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace('"', '')
 
    # Salvar as altera√ß√µes no arquivo Excel
    workbook.save(caminho_arquivo)
 
    # Fechar o arquivo Excel
    workbook.close()

def combinacoes_instancias(caminho_arquivo, aba):
    # Carregar a planilha Excel
    df = pd.read_excel(caminho_arquivo, sheet_name=aba)
 
    # Concatenar 'Group', 'Description' e 'Operating System' em uma nova coluna 'COMBINACAO'
    df['COMBINACAO'] = df['Group'] + ';' + df['Description'] + ';' + df['Operating System']+ ';' + df['Instance Type']
    df['STORAGE'] = df['Storage amount per Instance (GB)']
 
    # Calcular o maior "storage" para cada combina√ß√£o
    max_storage_combinacao = {}
    for _, row in df.iterrows():
        combinacao = row['COMBINACAO']
        storage = row['STORAGE']
        if combinacao not in max_storage_combinacao:
            max_storage_combinacao[combinacao] = storage
        else:
            max_storage_combinacao[combinacao] = max(max_storage_combinacao[combinacao], storage)
 
    # Contar o n√∫mero de ocorr√™ncias de cada combina√ß√£o
    contagem_combinacoes = df['COMBINACAO'].value_counts().reset_index()
    contagem_combinacoes.columns = ['COMBINACAO', 'QUANTIDADE']

    # Formatar os resultados no formato desejado e aplica calculo de Multi-AZ e Blue/Green
    def formatar_resultado(row):
        combinacao = row['COMBINACAO']
        quantidade = row['QUANTIDADE']
        partes = combinacao.split(';')
        max_storage = max_storage_combinacao.get(combinacao, 0)
        if len(partes) >= 3:
            grupo = partes[0]
            descricao = partes[1]
            so = partes[2]
            instancia = partes[3]
 
            # Obter a quantidade de m√°quinas
            qtde_maquinas = re.findall(r'(\d+) x', descricao)
            qtde_maquinas = int(qtde_maquinas[0])
 
            # Verificar se existem combina√ß√µes
            if quantidade > 1:
                nova_qtde_maquinas = qtde_maquinas + quantidade
                descricao = descricao.replace(f"{qtde_maquinas} x", f"{nova_qtde_maquinas} x")
 
                return f"{grupo};{nova_qtde_maquinas};{descricao};{so};{instancia};{max_storage}"
            else:
                return f"{grupo};{qtde_maquinas};{descricao};{so};{instancia};{max_storage}"
        else:
            return f"{combinacao};{quantidade};{max_storage}"
 
    contagem_combinacoes['RESULTADO'] = contagem_combinacoes.apply(formatar_resultado, axis=1)
 
    return contagem_combinacoes['RESULTADO']

def criar_aba_com_resultados_tratados(caminho_arquivo, resultados):
    workbook = load_workbook(caminho_arquivo)
    aba = 'Controle'
 
    # Deletar aba antiga
    if aba in workbook.sheetnames:
        antiga = workbook[aba]
        workbook.remove(antiga)
 
    # Criar nova aba de resultados
    nova_aba = workbook.create_sheet(aba)
 
    aba_controle = workbook[aba]
 
    mapeamento_colunas = {
        "Group": 0, "Description": 1, "AWS Region": 2, "Operating System": 3, "Instance Type": 4,
        "Tenancy": 5, "Number of Instances": 6, "Assumed Usage": 7, "Usage Type": 8, "Purchasing Option": 9,
        "Storage Type": 10, "Storage amount per Instance (GB)": 11,"Provisioning IOPS per instance (applicable for gp3, io1, io2)": 12, "EBS Throughput per Instance (applicable for gp3)(Mbps)": 13,
        "Snapshot Frequency": 14, "EBS Snapshot amount per Instance (GB/snapshot)": 15
    }
 
    # Adicionar cabe√ßalhos
    cabecalhos = list(mapeamento_colunas.keys())
    nova_aba.append(cabecalhos)
 
    ambiente_adicional = []
    # Preencher a nova aba com os resultados
    for resultado in resultados:
        if resultado is not None: #Verifica se o resultado n√£o √© None
            partes = resultado.split(';')
            descricao = '"' + partes[2] + '"'
 
            if partes[3] == "Linux Server":
                partes[3] = "Linux"
 
            if partes[0].upper() == "DEVELOPMENT" or partes[0].upper() == "DESENVOLVIMENTO" or partes[0].upper() == "HOMOLOGATION" or partes[0].upper() == "HOMOLOGACAO" or partes[0].upper() == "PRE-PRODUCTION" or partes[0].upper() == "PRE-PRODUCAO":
                linha = [partes[0], descricao,'sa-east-1', partes[3], partes[4], 'Shared Instances', partes[1], '40', 'Hours/Week', 'On-Demand', 'General Purpose SSD (gp3)', partes[5], '', '', '2x Daily', '10']
            elif partes[0].upper() == "PRODUCTION" or partes[0].upper() == "PRODUCAO":
                linha = [partes[0], descricao,'sa-east-1', partes[3], partes[4], 'Shared Instances', partes[1], '', 'Always On', '1 Yr No Upfront EC2 Instance Savings Plan', 'General Purpose SSD (gp3)', partes[5], '', '', '6x Daily', '20']
            else: # Outros ambientes
                linha = [partes[0], descricao,'sa-east-1', partes[3], partes[4], 'Shared Instances', partes[1], '40', 'Hours/Week', 'On-Demand', 'General Purpose SSD (gp3)', partes[5], '', '', '2x Daily', '10']
                ambiente_adicional.append(partes[0])
            nova_aba.append(linha)
 
    # Iterar sobre as c√©lulas da coluna B (Descricao) e remover as aspas
    for linha in aba_controle.iter_rows(min_row=2, min_col=2, max_col=2):  # Coluna B (Descricao) come√ßando da segunda linha
        for cell in linha:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace('"', '')
 
    # Salvar as altera√ß√µes no arquivo Excel
    workbook.save(caminho_arquivo)
 
    # Fechar o arquivo Excel
    workbook.close()

############################################################
############ RECOMENDA√á√ÉO DE INSTANCE TYPES ################
############################################################

def read_instance_types_from_txt(directory_path, file_name):
    file_path = os.path.join(directory_path, file_name)
    with open(file_path, "r") as file:
        instance_types = []
        for line in file:
            instance_info = line.strip().split(',')
            if len(instance_info) == 3:
                instance_types.append({
                    'InstanceType': instance_info[0],
                    'VCpuInfo': {'DefaultVCpus': int(instance_info[1])},
                    'MemoryInfo': {'SizeInMiB': float(instance_info[2]) * 1024}
                })
    # Ordenar os tipos de inst√¢ncia na ordem desejada: m6, m5, c6, c5
    instance_types_sorted = sorted(instance_types, key=lambda x: (
        x['InstanceType'].startswith('m6'),
        x['InstanceType'].startswith('m5'),
        x['InstanceType'].startswith('c6'),
        x['InstanceType'].startswith('c5')
    ), reverse=True)
    return instance_types_sorted

def extract_ram_and_vcpu(description):
    ram = re.findall(r'(\d+)GB RAM', description)
    vcpu = re.findall(r'(\d+)vCPU', description)
    if ram and vcpu:
        return int(ram[0]), int(vcpu[0])
    else:
        return None, None

def replace_ram_and_vcpu(description, info_ram, info_cpu):
    ram = re.findall(r'(\d+)GB RAM', description)
    vcpu = re.findall(r'(\d+)vCPU', description)
    if ram and vcpu:
        old_ram = f"{ram[0]}GB RAM"
        old_cpu = f"{vcpu[0]}vCPU"
 
        description = description.replace(old_ram, info_ram)
        description = description.replace(old_cpu, info_cpu)
 
        return description
    else:
        return None, None

def recommend_instance_type(vcpu_needed, ram_needed, instance_types):
    recommended_instance = None
    min_cost = float('inf')
    for instance_type in instance_types:
        vcpu_available = instance_type['VCpuInfo']['DefaultVCpus']
        ram_available = instance_type['MemoryInfo']['SizeInMiB'] / 1024
        if vcpu_needed <= vcpu_available and ram_needed <= ram_available:
            cost = ram_available
            if cost < min_cost:
                min_cost = cost
                ram = int(round(ram_available, 0))
                recommended_instance = f"{instance_type['InstanceType']},{ram}GB RAM,{vcpu_available}vCPU"
    return recommended_instance

############################################################
################### ARQUIVO TEMPLATE #######################
############################################################

def ocultar_aba(nome_arquivo, nome_aba):
    # Carregar o arquivo Excel
    wb = load_workbook(nome_arquivo)
 
    # Verificar se a aba existe
    if nome_aba in wb.sheetnames:
        # Ocultar a aba
        ws = wb[nome_aba]
        ws.sheet_state = 'hidden'
 
        # Salvar as altera√ß√µes
        wb.save(nome_arquivo)
    else:
        print(f'A aba "{nome_aba}" n√£o foi encontrada no arquivo.')

def copiar_arquivo(origem, destino):
    try:
        shutil.copy(origem, destino)
    except IOError as e:
        print(f"Erro ao copiar arquivo: {e}")

def selecionar_arquivo_Template():
    caminho_pasta = "data/template"
    arquivos = os.listdir(caminho_pasta)
    for arquivo in arquivos:
        arquivo_origem = os.path.join(os.getcwd(),caminho_pasta,arquivo)
    return arquivo_origem

def selecionar_destino_Template(sigla):
    caminho_pasta = "siglas"
    arquivo = sigla + "_AWS.xlsx"
    destino = os.path.join(os.getcwd(),caminho_pasta,arquivo)
    os.remove(destino) if os.path.exists(destino) else None
    return destino

def copiar_dados(origem, destino):
    # Carregar o arquivo Excel
    wb = load_workbook(origem)
    ws_origem = wb['Controle']
 
    wb = load_workbook(destino)
    ws_destino = wb['Inputs']
 
    # Limpar dados anteriores na aba de destino a partir da linha 4
    for row in ws_destino.iter_rows(min_row=4, min_col=2, max_col=ws_destino.max_column, max_row=ws_destino.max_row):
        for cell in row:
            cell.value = None
 
    arquivos_copiados = 0
    # Copiar dados da aba "Controle" para a aba de destino
    for row_idx, row in enumerate(ws_origem.iter_rows(min_row=2), start=4):
        if row[0].value.upper() in ["DEVELOPMENT", "DESENVOLVIMENTO", "HOMOLOGATION", "HOMOLOGACAO", "PRE-PRODUCTION", "PRE-PRODUCAO", "PRODUCTION", "PRODUCAO"]:
            arquivos_copiados = arquivos_copiados + 1
            for col_idx, cell in enumerate(row, start=2): # Come√ßa da coluna B
                ws_destino.cell(row=row_idx, column=col_idx).value = cell.value
 
    # Salvar as altera√ß√µes no arquivo de destino
    wb.save(destino)
 
    # Verificar se os dados foram copiados
    if arquivos_copiados == 0:
        os.remove(destino) if os.path.exists(destino) else None
        template_error_message = "<p style='font-size: 14px;'>Informa√ß√µes insuficientes para gerar o Modelo AWS ‚õî"
        show_warning_message("Aviso", template_error_message)
        exit(1)

############################################################
######################### EXEC #############################
############################################################

try:
    if sigla:
        caminho_arquivo = selecionar_arquivo()

        if caminho_arquivo:
            class WorkerThread(QThread):
                update_progress = pyqtSignal(int)

                def __init__(self, caminho_arquivo, sigla):
                    super().__init__()
                    self.caminho_arquivo = caminho_arquivo
                    self.sigla = sigla

                def run(self):
                    aba = "BASELINE"
                    if validar_aba_excel(self.caminho_arquivo, aba):
                        planilha = pd.read_excel(self.caminho_arquivo, sheet_name=aba)
                    else:
                        aba = 0
                        planilha = pd.read_excel(self.caminho_arquivo, sheet_name=aba)

                    total_linhas = len(planilha)
                    linhas_filtradas_siglas = []
                    for indice, valores in enumerate(planilha['SIGLAS'], start=1):
                        valores_split = str(valores).split(',')
                        for valor in valores_split:
                            if valor.strip() == self.sigla:
                                linhas_filtradas_siglas.append(indice - 1)
                        progresso = math.ceil((indice / total_linhas) * 100)
                        self.update_progress.emit(progresso)
                        time.sleep(0.1)
                    
                    global dados_filtrados_siglas
                    dados_filtrados_siglas = planilha.iloc[linhas_filtradas_siglas]
                    
                    window.close()

            class MainWindow(QMainWindow):
                def __init__(self):
                    super().__init__()
                    self.setWindowTitle("Carregando...")
                    self.setWindowIcon(QIcon(os.getcwd() + "/assets/ntt_icone.ico"))
                    self.setGeometry(800, 500, 400, 100)

                    self.layout = QVBoxLayout()
                    self.progress_bar = QProgressBar()
                    self.layout.addWidget(self.progress_bar)

                    self.central_widget = QWidget()
                    self.central_widget.setLayout(self.layout)
                    self.setCentralWidget(self.central_widget)

                    self.progress_bar.setValue(0)

                    self.worker_thread = None

                    self.start_progress()

                def update_progress_bar(self, value):
                    self.progress_bar.setValue(value)
                    if value == 100:
                        print("Carregamento conclu√≠do. \n")

                def start_progress(self):
                    self.worker_thread = WorkerThread(caminho_arquivo, sigla)
                    self.worker_thread.update_progress.connect(self.update_progress_bar)
                    self.worker_thread.start()

            window = MainWindow()
            window.show()
            app.exec_()
                
            #Continua√ß√£o do tratamento de dados
            qtde_siglas = dados_filtrados_siglas.__len__()

            if qtde_siglas == 0:
                sigla_no_message = f"<p style='font-size: 14px;'>Sigla <b style='color: red;'>{sigla}</b> n√£o encontrada"
                show_error_message("Error", sigla_no_message)
                exit(1)

            # Define os filtros adicionais desejados
            filtro_power_state = 'Powered On'
            filtro_tipo_aplicacao = 'BANCO DE DADOS'
            soWindows = False
            soLinux = False

            # Aplica os filtros adicionais nas colunas "PowerState" e "TIPO_APLICACAO"
            dados_filtrados = dados_filtrados_siglas[(dados_filtrados_siglas['PowerState'] == filtro_power_state) & (dados_filtrados_siglas['TIPO_APLICACAO'] != filtro_tipo_aplicacao)]

            # Aplica filtro adicional na coluna "CLASSE"
            if so_selecionado == "Windows":
                dados_filtrados = dados_filtrados[(dados_filtrados['CLASSE'] == 'Windows Server')]
                soPesquisa = "apenas Windows"
                if len(dados_filtrados) == 0:
                    so_win_no_message = f"<p style='font-size: 14px;'>A sigla <b style='color: red;'>{sigla}</b> n√£o possui servidores Windows"
                    show_error_message("Error", so_win_no_message)
                    exit(1)

            elif so_selecionado == "Linux":
                dados_filtrados = dados_filtrados[(dados_filtrados['CLASSE'] == 'Linux Server')]
                soPesquisa = "apenas Linux"
                if len(dados_filtrados) == 0:
                    so_linux_no_message = f"<p style='font-size: 14px;'>A sigla <b style='color: red;'>{sigla}</b> n√£o possui servidores Linux"
                    show_error_message("Error", so_linux_no_message)
                    exit(1)

            if so_selecionado == "Windows e Linux":
                if not dados_filtrados[(dados_filtrados['CLASSE'] == 'Windows Server')].empty:
                    soWindows = True
                if not dados_filtrados[(dados_filtrados['CLASSE'] == 'Linux Server')].empty:
                    soLinux = True
                if soWindows == True and soLinux == True:
                    soPesquisa = "Windows e Linux"
                elif soWindows == True and soLinux == False:
                    soPesquisa = "apenas Windows"
                elif soWindows == False and soLinux == True:
                    soPesquisa = "apenas Linux"

            # Escolha as colunas que deseja pegar os dados
            colunas_selecionadas = ['NOME','AMBIENTE','TIPO_APLICACAO','MEMORIA_RAM','vcpu','STORAGE','FUNCAO','CLASSE']

            # Obt√©m os dados das colunas selecionadas ap√≥s os filtros
            dados_selecionados = dados_filtrados[colunas_selecionadas]
            dados_selecionados = dados_selecionados.drop_duplicates(subset=["NOME"], keep="first")
            dados_selecionados.loc[:, "vcpu"] = dados_selecionados["vcpu"].fillna(2)
            dados_selecionados.loc[:, "vcpu"] = dados_selecionados["vcpu"].replace('', 2)
            dados_selecionados.loc[:, "vcpu"] = dados_selecionados["vcpu"].replace(0, 2)
            dados_selecionados.loc[:, "MEMORIA_RAM"] = dados_selecionados["MEMORIA_RAM"].fillna(4096)
            dados_selecionados.loc[:, "MEMORIA_RAM"] = dados_selecionados["MEMORIA_RAM"].replace('', 4096)
            dados_selecionados.loc[:, "MEMORIA_RAM"] = dados_selecionados["MEMORIA_RAM"].replace(0, 4096)
            dados_selecionados.loc[:, "MEMORIA_RAM"] /= 1024
            dados_selecionados.loc[:, "MEMORIA_RAM"] = dados_selecionados.loc[:, "MEMORIA_RAM"].apply(lambda x: math.ceil(x))
            dados_selecionados.loc[:, "STORAGE"] = dados_selecionados["STORAGE"].fillna(150)
            dados_selecionados.loc[:, "STORAGE"] = dados_selecionados["STORAGE"].replace('', 150)
            dados_selecionados.loc[:, "STORAGE"] = dados_selecionados["STORAGE"].apply(lambda x: math.ceil(x))

            if (dados_selecionados["STORAGE"] == 0).any():
                storage_no_message = f"<p style='font-size: 14px; color: red; font-weight: bold;'>Aten√ß√£o!"
                storage_no_message += f"<p style='font-size: 14px;'>H√° servidores com STORAGE = 0 (zero) na sigla: <b style='color: red;'>{sigla}</b>"
                storage_no_message += f"<p style='font-size: 14px;'>Por favor verifique no CMDB a quantidade de STORAGE correta do servidor"
                storage_no_message += "<p style='font-size: 14px;'><a href='https://itau.service-now.com/now/nav/ui/classic/params/target/cmdb_ci_server_list.do%3Fsysparm_userpref_module'>Clique aqui para acessar o CMDB</a>"
                show_information_message("Aviso", storage_no_message)
                print(f"\n H√° servidores com dados == 0 (zero) no STORAGE da sigla: {sigla}. Por favor verifique no CMDB a quantidade de STORAGE do servidor. \n \nLink do CMDB: https://itau.service-now.com/now/nav/ui/classic/params/target/cmdb_ci_server_list.do%3Fsysparm_userpref_module")

            # Cria uma nova planilha com os dados selecionados
            novo_caminho_arquivo = selecionar_caminho_salvar()

            # Verifica se o usu√°rio selecionou um local para salvar o arquivo
            if novo_caminho_arquivo:
                # Salva os dados no arquivo criado
                dados_selecionados.to_excel(novo_caminho_arquivo, index=False)

                # Abre o arquivo Excel criado
                workbook = load_workbook(novo_caminho_arquivo)

                # Altera o nome da primeira aba
                primeira_aba = workbook.sheetnames[0]  # Pega o nome da primeira aba
                novo_nome_aba = "BASELINE_SIGLA"
                workbook[primeira_aba].title = novo_nome_aba

                # Salva as altera√ß√µes no arquivo Excel
                workbook.save(novo_caminho_arquivo)

                # Converte caracteres em ISO-8859-1 para UTF-8
                converter_utf8(novo_caminho_arquivo, novo_nome_aba)

                # Renomear as colunas e ordenar
                organizar_colunas(novo_caminho_arquivo, novo_nome_aba)

                # Verifica as combina√ß√µes
                resultados = combinacoes_configuracoes(novo_caminho_arquivo, novo_nome_aba)

                # Cria nova aba com os dados tratados
                criar_aba_com_resultados(novo_caminho_arquivo, resultados)

                # Ler a aba "Controle"
                abaControle = 'Controle'
                data_controle = pd.read_excel(novo_caminho_arquivo, sheet_name=abaControle)

                # Ler os tipos de inst√¢ncia do arquivo txt
                directory_path = "data/servers"
                file_name = "instance_types.txt"
                instance_types = read_instance_types_from_txt(directory_path, file_name)

                workbook = load_workbook(novo_caminho_arquivo)
                worksheet = workbook[abaControle]

                for index, row in data_controle.iterrows():
                    description = row['Description']
                    ram_needed, vcpu_needed = extract_ram_and_vcpu(description)
                    if ram_needed is not None and vcpu_needed is not None:
                        recommended_instance_type = recommend_instance_type(vcpu_needed, ram_needed, instance_types)
                        info_configuration = recommended_instance_type.split(',')

                        # Defini√ß√µes do instance type
                        instance_type = info_configuration[0]
                        info_ram = info_configuration[1]
                        info_cpu = info_configuration[2]

                        # Instance type
                        worksheet.cell(row=index + 2, column=5, value=instance_type)

                        # Nova descri√ß√£o
                        new_description = replace_ram_and_vcpu(description, info_ram, info_cpu)
                        worksheet.cell(row=index + 2, column=2, value=new_description)

                workbook.save(novo_caminho_arquivo)

                ocultar_aba(novo_caminho_arquivo, novo_nome_aba)

                # Verifica as combina√ß√µes
                dados_tratados = combinacoes_instancias(novo_caminho_arquivo, "Controle")

                # Cria nova aba com os dados tratados
                criar_aba_com_resultados_tratados(novo_caminho_arquivo, dados_tratados)

                arquivoTemplateAWS = selecionar_arquivo_Template()
                if arquivoTemplateAWS:
                    destinoModeloAWS = selecionar_destino_Template(sigla)
                    if destinoModeloAWS:
                        copiar_arquivo(arquivoTemplateAWS, destinoModeloAWS)
                    else:
                        print("Nenhum destino selecionado.")
                else:
                    print("Nenhum arquivo selecionado.")

                # Copia os dados para o modelo AWS
                copiar_dados(novo_caminho_arquivo, destinoModeloAWS)

                print(f"Tratamento de dados conclu√≠do com sucesso. \n\nOs servidores s√£o {soPesquisa}.")

                print("\n ------------------------------------------------------------ \n")
        else:
            print("Nenhum arquivo selecionado. A a√ß√£o foi cancelada pelo usu√°rio.")
            exit(1)
except NameError:
    print("Calculadora encerrada pelo usu√°rio.")